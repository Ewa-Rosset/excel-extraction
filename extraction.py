import openpyxl
from openpyxl import Workbook
import PySimpleGUI as sg
import time
import os

""" Ask user to input the relevant POE file """

sg.theme('Green Mono') 

layout = [[sg.Text('Upload a POE file for reformatting. Program will return a new file')],
      [sg.Text('Source for file', size=(15, 1)), sg.InputText(), sg.FileBrowse()],
      [sg.Submit(), sg.Cancel()]]

window = sg.Window('Proof of Eligibility Reformatting', layout)

event, values = window.read()
window.close()
file_path = values[0]    



""" Opening the original spreadsheet"""

wb1 = openpyxl.load_workbook(file_path)


"""Selecting the relevant sheet we will be working on and giving it a variable name"""

eligibility = wb1["eligibility"]


""" Creating a new workbook with a sheet name called Eligibility2 """

new_wb = openpyxl.Workbook()
eligibility2 = new_wb.active


eligibility2.title = 'Eligibility2'
#print(eligibility2.title)


""" Add the relevant two columns from the previous sheet """

num_of_cols = eligibility.max_column
num_of_rows = eligibility.max_row

for row_num in range(1, (num_of_rows+1)):
     eligibility2.cell(row=row_num, column=1).value = eligibility.cell(row=row_num, column=2).value
     eligibility2.cell(row=row_num, column=2).value = eligibility.cell(row=row_num, column=4).value


file_name = os.path.basename(file_path)

ts = str(time.time())

newtitle = "Updated" + ts + file_name

print("newtitle: ", newtitle)

new_wb.save(newtitle)